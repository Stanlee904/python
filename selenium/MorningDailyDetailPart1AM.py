from datetime import datetime
from openpyxl import load_workbook
from pywinauto import application
import time
import pyautogui
import pandas
import os


part1AMCountList = {"BNK투자증권(INDEX DATA/INFO)" : 0,
                    "BNK투자증권   (VOL 데이터)" : 0,
                    "BNK투자증권   (환율/스왑금리)" : 0,
                    "KTB투자증권   (LIBOR 금리10AM)" : 0,
                    "한화손해보험  ( 환율 )": 0,
                    "흥국생명(환율) 9시30분이후": 0,
                    "KTB투자증권   (LIBOR 금리)" : 0}

curTime = datetime.today()
year = curTime.strftime("%Y")
month = curTime.strftime("%m")
day = curTime.strftime("%d")


logProgramPath = 'D:\SVN\VB6_추가사항\신규추가사항.exe'
wbPart1CheckFile = 'D:\workspace\python\selenium\PART1AM확인용.xlsx'
wbPart1FilterStateFile = 'D:\workspace\python\selenium\PART1AMFilter상태.xlsx'
wbPart1FilterFile = 'D:\workspace\python\selenium\PART1AM필터적용.xlsx'




try:
    def ChkPart1AM(dailySaveFile):
                
        #1. 신규추가사항 Open
        app = application.Application(backend='uia').start(logProgramPath)
        
        #2. App 접근 
        dlg = app['전송모니터 프로그램 - 2022-09-30 #1 전송상태로그 엑셀파일 다운로드 시, 예외처리 추가']
        
        #3. 메뉴바(쿼터진행상태관리) 클릭
        dlg.child_window(title="쿼터진행상태관리", control_type="MenuItem").select()
        
        #4. 쿼터진행상태관리 -> 전송상태확인 클릭
        dlg['전송상태확인MenuItem2'].click_input()
        
        #5. 조회 버튼 클릭 
        dlg['Button9'].click_input()
        
        #6. DOWN 버튼 클릭
        dlg['DOWNButton'].click_input()        

        #7. Warning 창으로 인해 타임슬립 추가
        time.sleep(10)
        
        #8. 저장버튼 keyPress로 진행
        pyautogui.press('s')

        #9. 다른 이름으로 저장 화면보다 코드가 빨리 진행 될 수 있기 때문에 타임슬립 추가
        time.sleep(10)        
        
        #10. 저장되지 않은 new file을 App api로 접근함. 
        app2 = application.Application(backend='uia').connect(title_re="통합 문서1 - Excel")
        
        #11. dialog 접근
        dlg2 = app2['통합 문서1 - Excel']       
        
        #12. 경로 작성
        dlg2['파일 이름:Edit'].type_keys(wbPart1CheckFile) 
        
        #12. 저장 버튼 클릭
        dlg2['저장(S)'].click()
        
        time.sleep(2)
        
        #신규추가사항 종료
        app.kill()
        
        #엑셀 불러오기
        wbPart1Check = load_workbook(wbPart1CheckFile)
        
        #시트 접근
        workSheetPart1Check = wbPart1Check["Sheet1"]
        
        #필터 설정
        workSheetPart1Check.auto_filter.ref = workSheetPart1Check.dimensions
        
        #컬럼 필터 추가
        workSheetPart1Check.auto_filter.add_filter_column(1, ['Part 1 AM'])        
        
        #다른이름으로 저장
        wbPart1Check.save(wbPart1FilterStateFile)
        
        #pandas로 엑셀 불러오기(왜냐하면 openpyxl은 필터가 설정만 되고 실질적인 필터는 안됨. )
        dataFrame = pandas.read_excel(wbPart1FilterStateFile, sheet_name="Sheet1")
        
        #필터 조건
        dataFilter = dataFrame.loc[(dataFrame['PART'] == 'Part 1 AM') & (dataFrame['로그타입'] == 'FTP전송완료') ]

        #다른이름으로 저장
        dataFilter.to_excel(wbPart1FilterFile)        
        
        #엑셀 불러오기
        wbPart1FilterApply = load_workbook(wbPart1FilterFile)
        
        #엑셀 시트 접근
        workSheetPart1FilterApply = wbPart1FilterApply["Sheet1"]

        #D컬럼 전체 값 가져오기
        column_D = workSheetPart1FilterApply["D"]
        
        #값 비교해서 FTP전송완료 된 부분만 COUNT 하기
        if len(column_D) > 0:           
            for index in range (2,len(column_D)+1):
                if workSheetPart1FilterApply['D'+str(index)].value == 'BNK투자증권(INDEX DATA/INFO)':
                    part1AMCountList['BNK투자증권(INDEX DATA/INFO)'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == 'BNK투자증권   (VOL 데이터)':
                    part1AMCountList['BNK투자증권   (VOL 데이터)'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == 'BNK투자증권   (환율/스왑금리)':
                    part1AMCountList['BNK투자증권   (환율/스왑금리)'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == 'KTB투자증권   (LIBOR 금리10AM)':
                    part1AMCountList['KTB투자증권   (LIBOR 금리10AM)'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == '한화손해보험  ( 환율 )':
                    part1AMCountList['한화손해보험  ( 환율 )'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == '흥국생명(환율) 9시30분이후':
                    part1AMCountList['흥국생명(환율) 9시30분이후'] += 1
                elif workSheetPart1FilterApply['D'+str(index)].value == 'KTB투자증권   (LIBOR 금리)':
                    part1AMCountList['KTB투자증권   (LIBOR 금리)'] += 1                
        
        #데일리 엑셀 불러오기
        wbDaily = load_workbook(dailySaveFile)
        
        #엑셀 시트 접근
        workSheetDaily = wbDaily["Sheet1"]

        #A컬럼 값 가져오기
        wsColumnA = workSheetDaily["A"]
        
        # 값 비교하여 O값 넣기
        if len(wsColumnA) > 0:
            for index2 in range(18,len(wsColumnA)+1):
                if workSheetDaily['N'+str(index2)].value == part1AMCountList[workSheetDaily['A'+str(index2)].value]:
                    workSheetDaily['M'+str(index2)] = 'O'        
         
        #저장
        wbDaily.save(dailySaveFile)
        
        os.remove(wbPart1CheckFile)
        os.remove(wbPart1FilterStateFile)
        os.remove(wbPart1FilterFile)
        
        
except Exception as ex:
    print(ex)