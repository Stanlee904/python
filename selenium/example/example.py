# import os 
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from pywinauto import application
from pywinauto import findwindows
from openpyxl import load_workbook

import time
import pyautogui

# import pandas as pd
# import numpy as np
# import re


# file_path = "D:\새 폴더\오전데일리 자동화_20230821.xlsx"


# if os.path.exists(file_path):
#     os.remove(file_path)
# else:
#     print("파일 존재하지 않음")



curTime = datetime.today()
year = curTime.strftime("%Y")
month = curTime.strftime("%m")
day = curTime.strftime("%d")
hour = curTime.strftime("%H")
minute = curTime.strftime("%M")
second = curTime.strftime("%S")


# if int(hour) >= 15:
#     print("okay")
# else:
#     print("not okay")

# procs = findwindows.find_elements()

# for proc in procs:
#      print(f"{proc}  / 프로세스: {proc.process_id}")


# app = application.Application(backend='uia').start("D:\\SVN\VB6_추가사항\\신규추가사항.exe")
# dlg = app['전송모니터 프로그램 - 2022-09-30 #1 전송상태로그 엑셀파일 다운로드 시, 예외처리 추가']
# # dlg.child_window(title="쿼터진행상태관리", control_type="MenuItem").click_input() 메뉴바 클릭


# # 서브 메뉴 선택방법 => 메인메뉴 선택 후 -> dlg.print_control_identifiers()로 서브메뉴 정보 확인
# # dlg.child_window(title="쿼터진행상태관리", control_type="MenuItem").select()
# # dlg.print_control_identifiers()

# dlg.child_window(title="쿼터진행상태관리", control_type="MenuItem").select()
# dlg['전송상태확인MenuItem2'].click_input()
# dlg['Button9'].click_input()
# dlg['DOWNButton'].click_input()


# time.sleep(5)

# # pyautogui.press('esc')

# app2 = application.Application(backend='uia').connect(title_re="통합 문서1 - Excel")
# dlg2 = app['통합 문서1 - Excel']

# pyautogui.press('s')

# time.sleep(5)

# # app2 = application.Application(backend='uia').connect(title_re="다른 이름으로 저장")
# # dlg2 = app2['다른 이름으로 저장']

# app2 = application.Application(backend='uia').connect(title_re="통합 문서1 - Excel")
# dlg2 = app2['통합 문서1 - Excel']



# dlg2['파일 이름:Edit'].type_keys('D:\\workspace\\python\\selenium\\PART1AM확인용.xlsx')

# # dlg2.child_window(title="저장(S)", auto_id="1", control_type="Button").select()

# dlg2['저장(S)'].click()



# time.sleep(2)


# app.kill()


# wb = load_workbook('D:\workspace\python\selenium\PART1AM확인용.xlsx')
# workSheet = wb["Sheet1"]

# workSheet.auto_filter.ref = workSheet.dimensions
# workSheet.auto_filter.add_filter_column(1, ['Part 1 AM'])

# wb.save('D:\workspace\python\selenium\PART1AM확인용2.xlsx')

# df = pd.read_excel('D:\workspace\python\selenium\PART1AM확인용2.xlsx', sheet_name="Sheet1")
# df1 = df.loc[(df['PART'] == 'Part 1 AM') & (df['로그타입'] == 'FTP전송완료') ]


# df1.to_excel("D:\workspace\python\selenium\PART1AM필터적용.xlsx")



# wb2 = load_workbook('D:\workspace\python\selenium\PART1AM필터적용.xlsx')
# workSheet2 = wb2["Sheet1"]

# column_D = workSheet2["D"]

# part1AMCountList = {"BNK투자증권(INDEX DATA/INFO)" : 0,
#                     "BNK투자증권   (VOL 데이터)" : 0,
#                     "BNK투자증권   (환율/스왑금리)" : 0,
#                     "KTB투자증권   (LIBOR 금리10AM)" : 0,
#                     "한화손해보험  ( 환율 )": 0,
#                     "흥국생명(환율) 9시30분이후": 0,
#                     "KTB투자증권   (LIBOR 금리)" : 0}

# if len(column_D) > 0:
    
#     for index in range (2,len(column_D)+1):
        
#         if workSheet2['D'+str(index)].value == 'BNK투자증권(INDEX DATA/INFO)':
#             part1AMCountList['BNK투자증권(INDEX DATA/INFO)'] += 1
#         elif workSheet2['D'+str(index)].value == 'BNK투자증권   (VOL 데이터)':
#             part1AMCountList['BNK투자증권   (VOL 데이터)'] += 1
#         elif workSheet2['D'+str(index)].value == 'BNK투자증권   (환율/스왑금리)':
#             part1AMCountList['BNK투자증권   (환율/스왑금리)'] += 1
#         elif workSheet2['D'+str(index)].value == 'KTB투자증권   (LIBOR 금리10AM)':
#             part1AMCountList['KTB투자증권   (LIBOR 금리10AM)'] += 1
#         elif workSheet2['D'+str(index)].value == '한화손해보험  ( 환율 )':
#             part1AMCountList['한화손해보험  ( 환율 )'] += 1
#         elif workSheet2['D'+str(index)].value == '흥국생명(환율) 9시30분이후':
#             part1AMCountList['흥국생명(환율) 9시30분이후'] += 1
#         elif workSheet2['D'+str(index)].value == 'KTB투자증권   (LIBOR 금리)':
#             part1AMCountList['KTB투자증권   (LIBOR 금리)'] += 1

# print(part1AMCountList.values())

# wb3 = load_workbook('D:\새 폴더\오전데일리 자동화.xlsx')
# workSheet3 = wb3["Sheet1"]

# wsColumnA = workSheet3["A"]

# print(len(wsColumnA))

# if len(wsColumnA) > 0:
#     for index2 in range(15,len(wsColumnA)+1):
#         if workSheet3['N'+str(index2)].value == part1AMCountList[workSheet3['A'+str(index2)].value]:
#             workSheet3['M'+str(index2)] = 'O'

            
# wb3.save("D:\새 폴더\오전데일리 자동화_"+year+month+day+".xlsx")

try:
        
    app = application.Application(backend='uia').start("C:\\Users\\user\\AppData\\Roaming\\NICE P&I\\NICE P&I\\NICE피앤아이.exe")
    dlg = app['NICE피앤아이 V 2.81']
    
    dlg.child_window(title="통합시스템", auto_id="8", control_type="Button").click_input()
    
    time.sleep(5)
    
    procs = findwindows.find_elements()

    for proc in procs:
        # print(f"{proc}  / 프로세스: {proc.process_id}")
        tempProc = f"{proc}"        
        if '통합 System' in tempProc:
            tempProcessId = proc.process_id
        
    app2 = application.Application(backend='uia').connect(process=tempProcessId)
    
    # print("**************||" +str(tempProcessId) + "||****************")
    dlg2 = app2['Dialog']
    
    dlg2.child_window(title="기타", control_type="MenuItem").select()
            
    #time.sleep(3)
    
    dlg2['세금계산MenuItem2'].select()
        
    dlg2.print_control_identifiers()
    
    # dlg.child_window(title="10.0.1.45_배치_Win2012_Server_R2", auto_id="1448272", control_type="Button").click()
    # dlg['10.0.1.45_배치_Win2012_Server_R2TreeItem'].select()
    
    # time.sleep(4)
    #8. 저장버튼 keyPress로 진행
    # pyautogui.press('ENTER')
    
    # time.sleep(4)
    #8. 저장버튼 keyPress로 진행
    # pyautogui.press('ENTER')
    
    # dlg.print_control_identifiers()
    
    # time.sleep(5)
    
    # procs = findwindows.find_elements()

    # for proc in procs:
    #     print(f"{proc}  / 프로세스: {proc.process_id}")
    #     tempProc = f"{proc}"        
    #     if '통합 System' in tempProc:
    #         tempProcessId = proc.process_id
        
    # app2 = application.Application(backend='uia').connect(process=tempProcessId)
    
    # dlg2 = app2['Dialog']
    
    # dlg2.print_control_identifiers()
    
    # 해당 이미지가 존재 한다면 
    # if (pyautogui.locateOnScreen("D:\\workspace\\python\\selenium\\IMG\\NiceIntegrated2.png") is not None):
    #     print("O")
    #     # workSheet["M12"] = 'O'
    # else:
    #     print("asdf")
    #     # workSheet["M12"] = '통합시스템 내에 확인 요망(예탁원 오전 자료수신)'    
        
    # if (pyautogui.locateOnScreen("D:\\workspace\\python\\selenium\\IMG\\NiceIntegrated3.png") is not None):
    #     print("S")
    # else:        
    #     print("1234")
    
    
    
    # if (pyautogui.locateOnScreen("D:\\workspace\\python\\selenium\\IMG\\NiceIntegrated4.png") is not None):
    #     print("D")
    # else:
    #     print("1234444")


except Exception as ex:
    print(ex)
