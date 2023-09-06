from openpyxl import Workbook
from openpyxl import load_workbook

saveValue = "D:\팀업무교육자료/save.xlsx"

#엑셀 파일 로드 하기
wb = load_workbook(saveValue)

# 시트 활성화
workSheet = wb["First_Sheet"]

workSheet.cell(row = 1,column = 2, value = 30)
workSheet.cell(row = 1,column = 1, value = 40)

workSheet["A2"] = 50


print(workSheet["A2"].value)

# 셀 값 삭제
workSheet["A2"] = ""


#행 삭제
workSheet.delete_rows(1)

#열 삭제
workSheet.delete_cols(1)

#빈 행 추가
workSheet.insert_cols(2)
workSheet.insert_cols(1,3) # 1부터 3개 삽입

#빈 열 추가
workSheet.insert_rows(2)
workSheet.insert_rows(1,3) # 1부터 3개 삽입

# 입력된 내용을 그대로 다른곳으로 이동하는 경우
workSheet.move_range("D1", cols=2)

workSheet.move_range("D1:F10", rows=-1, cols=2)



wb.save(saveValue)


# wb = Workbook()
# # 시트 활성화
# workSheet = wb.active
# # 시트 추가
# workSheet_new = wb.create_sheet() 

# workSheet.title = "First_Sheet"
# workSheet_new.title = "New Sheet"


# ws_want = wb.create_sheet("Third sheet", 0)

# # 시트탭 색깔 변경
# workSheet.sheet_properties.tabColor = "f4566e"

# wb.save("D:\팀업무교육자료/save.xlsx")


